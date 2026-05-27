#!/usr/bin/env python3
"""Profile the ANSES/Ciqual dataset before importing it into the database.

This script does not modify the database. It only reads the source files
and prints useful information about their structure and data quality.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[3]

ANSES_DIR = PROJECT_ROOT / "temp" / "EuropeNutrientsDBs" / "anses"
MAIN_FILE = ANSES_DIR / "Table Ciqual 2025_ENG_2025_11_03.xlsx"
DOC_FILE = ANSES_DIR / "Table Ciqual 2025 doc ENG_2025_11_19.pdf"

FIRST_NUTRIENT_COLUMN = "Energy,\nRegulation\nEU No\n1169\n2011 (kJ\n100g)"
FOOD_CODE_COLUMN = "alim_code"
CATEGORY_COLUMN = "alim_grp_nom_eng"
SUBCATEGORY_COLUMN = "alim_ssgrp_nom_eng"
SUBSUBCATEGORY_COLUMN = "alim_ssssgrp_nom_eng"


def get_columns(worksheet) -> list[str]:
    """Return the Excel header."""
    first_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    return list(first_row)


def print_main_file_profile(
    worksheet,
    workbook_sheetnames: list[str],
    metadata_columns: list[str],
    nutrient_columns: list[str],
) -> None:
    """Print basic information for the main ANSES file."""
    print()
    print("=" * 30)
    print("ANSES/Ciqual file profile")
    print("=" * 30)
    print(f"Main file: {MAIN_FILE}")
    print(f"Documentation File: {DOC_FILE}")
    print()
    print("Sheet names of Main file")
    print("=" * 30)
    print(f"sheetnames: {workbook_sheetnames}")
    print()
    print(f"Sheet name: {worksheet.title}")
    print("=" * 30)
    print(f"Total rows: {worksheet.max_row}")
    print(f"Total columns: {worksheet.max_column}")
    print(f"Metadata columns: {len(metadata_columns)}")

    for column in metadata_columns:
        print(f"- {column}")

    print(f"Nutrient columns: {len(nutrient_columns)}")

    for column in nutrient_columns[:20]:
        print(f"- {column}")


def print_first_foods(
    worksheet,
    columns: list[str],
) -> None:
    """Print a small sample of ANSES foods."""

    foods = worksheet.iter_rows(min_row=2, max_row=6, values_only=True)

    for food in foods:
        food_row = dict(zip(columns, food))
        print(
            f"- {food_row['alim_code']} | "
            f"{food_row['alim_grp_nom_eng']} | "
            f"{food_row['alim_ssgrp_nom_eng']} | "
            f"{food_row['alim_ssssgrp_nom_eng']} | "
            f"{food_row['alim_nom_eng']}"
        )


def split_metadata_and_nutrient_columns(
    columns: list[str],
) -> tuple[list[str], list[str]]:
    """Split ANSES main columns into food metadata and nutrient columns."""

    first_nutrient_column_index = columns.index(FIRST_NUTRIENT_COLUMN)
    metadata_columns = columns[:first_nutrient_column_index]
    nutrient_columns = columns[first_nutrient_column_index:]

    return metadata_columns, nutrient_columns


def print_food_code_uniqueness(
    worksheet,
    columns: list[str],
) -> None:
    """Print whether ANSES food codes are unique."""
    codes: set[str] = set()
    duplicate_codes: list[str] = []
    rows = worksheet.iter_rows(min_row=2, values_only=True)

    for row in rows:
        food_row = dict(zip(columns, row))
        food_code = str(food_row[FOOD_CODE_COLUMN]).strip()

        if food_code in codes:
            duplicate_codes.append(food_code)
        else:
            codes.add(food_code)

    print("=" * 30)
    print(
        f"Unique codes: {len(codes)}\n"
        f"Duplicate codes: {len(duplicate_codes)}"
    )

    if duplicate_codes:
        print("First duplicate codes")
        for code in duplicate_codes[:20]:
            print(f"- {code}")
      
    print("=" * 30)
    print()


def print_category_profile(
    worksheet,
    columns: list[str],
) -> None:
    """Print category profile."""
    categories: set[str] = set()
    subcategories: set[str] = set()
    subsubcategories: set[str] = set()

    null_row_counter_categories: int = 0
    null_row_counter_subcategories: int = 0
    null_row_counter_subsubcategories: int = 0

    rows = worksheet.iter_rows(min_row=2, values_only=True)

    for row in rows:
        food_row = dict(zip(columns, row))
        category = str(food_row[CATEGORY_COLUMN]).strip()
        subcategory = str(food_row[SUBCATEGORY_COLUMN]).strip()
        subsubcategory = str(food_row[SUBSUBCATEGORY_COLUMN]).strip()

        if category == "-":
            null_row_counter_categories += 1
        else:
            categories.add(category)

        if subcategory == "-":
            null_row_counter_subcategories += 1
        else:
            subcategories.add(subcategory)

        if subsubcategory == "-":
            null_row_counter_subsubcategories += 1
        else:
            subsubcategories.add(subsubcategory)

    print("=" * 30)
    print(f"Categories: {len(categories)}")
    print(f"Subcategories: {len(subcategories)}")
    print(f"Subsubcategories: {len(subsubcategories)}\n")

    print(f"Count categories null rows: {null_row_counter_categories}")
    print(f"Count subcategories null rows: {null_row_counter_subcategories}")
    print(f"Count subsubcategories null rows: {null_row_counter_subsubcategories}\n")

    for category in sorted(categories)[:20]:
        print(f"Categories: - {category}")

    for subcategory in sorted(subcategories)[:20]:
        print(f"Subcategories: - {subcategory}")

    for subsubcategory in sorted(subsubcategories)[:20]:
        print(f"Subsubcategories: - {subsubcategory}")

    print("=" * 30)


def main() -> int:
    """Run the ANSES profiling checks."""
    if not MAIN_FILE.exists():
        print(f"File not found: {MAIN_FILE}")
        return 1

    if not DOC_FILE.exists():
        print(f"File not found: {DOC_FILE}")
        return 1

    workbook = load_workbook(MAIN_FILE, read_only=True, data_only=True)
    workbook_sheetnames = workbook.sheetnames

    worksheet = workbook["food composition"]

    columns = get_columns(worksheet)

    metadata_columns, nutrient_columns = split_metadata_and_nutrient_columns(columns)

    print_main_file_profile(
        worksheet,
        workbook_sheetnames,
        metadata_columns,
        nutrient_columns,
    )
    print_first_foods(worksheet, columns)
    print_food_code_uniqueness(worksheet, columns)
    print_category_profile(worksheet, columns)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())