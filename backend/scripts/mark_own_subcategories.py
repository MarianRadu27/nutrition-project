from __future__ import annotations

import argparse
import math
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


HEADER_ALIASES = {
    "dacode": "da_code",
    "dacodecode": "da_code",
    "fooddescription": "food_description",
    "ownsubcategory": "own_subcategory",
    "isownsubcategory": "own_subcategory",
}


def normalize_header(value: Any) -> str:
    """Normalize Excel headers so small formatting differences do not matter."""
    if value is None:
        return ""
    return "".join(character for character in str(value).lower() if character.isalnum())


def resolve_path(path_text: str) -> Path:
    """Resolve relative paths from the project root when the script is run there."""
    path = Path(path_text)
    if path.is_absolute():
        return path
    return (Path.cwd() / path).resolve()


def looks_like_da_code(value: Any) -> bool:
    """Return True only for real numeric food codes, not category heading text."""
    if value is None:
        return False
    if isinstance(value, bool):
        return False
    if isinstance(value, int):
        return True
    if isinstance(value, float):
        return not math.isnan(value)

    text = str(value).strip()
    if not text:
        return False
    return text.endswith(".0") and text[:-2].isdigit() or text.isdigit()


def build_column_lookup(header_values: list[Any]) -> dict[str, int]:
    """Map the required logical column names to Excel column indexes."""
    lookup: dict[str, int] = {}
    for index, value in enumerate(header_values, start=1):
        alias = HEADER_ALIASES.get(normalize_header(value))
        if alias:
            lookup[alias] = index
    return lookup


def find_or_create_own_subcategory_column(sheet) -> int:
    """Reuse an existing own_subcategory column or create one in the first empty header."""
    header_values = [cell.value for cell in sheet[1]]
    lookup = build_column_lookup(header_values)
    if "own_subcategory" in lookup:
        return lookup["own_subcategory"]

    for cell in sheet[1]:
        if cell.value is None:
            cell.value = "own_subcategory"
            return cell.column

    next_column = sheet.max_column + 1
    sheet.cell(row=1, column=next_column).value = "own_subcategory"
    return next_column


def mark_workbook(input_path: Path, output_path: Path, sheet_name: str | None) -> None:
    """Mark bold food rows that should become their own Food group."""
    workbook = load_workbook(input_path)
    sheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]

    header_values = [cell.value for cell in sheet[1]]
    lookup = build_column_lookup(header_values)
    missing_columns = {"da_code", "food_description"} - set(lookup)
    if missing_columns:
        missing_text = ", ".join(sorted(missing_columns))
        raise RuntimeError(f"Missing required columns: {missing_text}")

    da_code_column = lookup["da_code"]
    description_column = lookup["food_description"]
    own_subcategory_column = find_or_create_own_subcategory_column(sheet)

    food_rows = 0
    marked_rows = 0

    for row_index in range(2, sheet.max_row + 1):
        da_code_cell = sheet.cell(row=row_index, column=da_code_column)
        description_cell = sheet.cell(row=row_index, column=description_column)
        marker_cell = sheet.cell(row=row_index, column=own_subcategory_column)

        if not looks_like_da_code(da_code_cell.value):
            marker_cell.value = None
            continue

        food_rows += 1
        is_bold_food = bool(da_code_cell.font.bold or description_cell.font.bold)
        marker_cell.value = is_bold_food

        if is_bold_food:
            marked_rows += 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    print(f"Input:  {input_path}")
    print(f"Output: {output_path}")
    print(f"Sheet:  {sheet.title}")
    print(f"Food rows checked: {food_rows}")
    print(f"own_subcategory TRUE rows: {marked_rows}")


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Create an Excel copy with own_subcategory marked from bold food rows."
    )
    parser.add_argument(
        "--input",
        default=r"backend\data\FoodsFinalTest.xlsx",
        help="Excel file to read.",
    )
    parser.add_argument(
        "--output",
        default=r"backend\data\FoodsFinalTest_marked.xlsx",
        help="Excel file to create.",
    )
    parser.add_argument("--sheet", help="Sheet to process. Defaults to the first sheet.")
    return parser.parse_args()


def main() -> int:
    args = parse_arguments()
    input_path = resolve_path(args.input)
    output_path = resolve_path(args.output)

    if not input_path.exists():
        print(f"ERROR: Excel file not found: {input_path}")
        return 1

    mark_workbook(input_path=input_path, output_path=output_path, sheet_name=args.sheet)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
