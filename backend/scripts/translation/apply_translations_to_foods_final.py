from __future__ import annotations

"""Copy reviewed translations into a FoodsFinal Excel copy for central storage."""

import argparse
import math
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[3]
DEFAULT_REVIEW = PROJECT_ROOT / "backend" / "data" / "translation_review_with_suggestions.xlsx"
DEFAULT_INPUT = PROJECT_ROOT / "backend" / "data" / "FoodsFinal.xlsx"
DEFAULT_OUTPUT = PROJECT_ROOT / "backend" / "data" / "FoodsFinal_with_ro.xlsx"


def resolve_path(path_text: str) -> Path:
    """Resolve paths from the project root unless an absolute path is provided."""
    path = Path(path_text)
    if path.is_absolute():
        return path
    return (PROJECT_ROOT / path).resolve()


def is_blank(value: Any) -> bool:
    """Treat pandas/openpyxl missing values and empty strings as blank."""
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    return str(value).strip() == ""


def clean_text(value: Any) -> str:
    """Convert values to clean strings."""
    if is_blank(value):
        return ""
    return str(value).strip()


def clean_int(value: Any) -> int | None:
    """Read integer identifiers that may come from Excel as floats."""
    if is_blank(value):
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        return None

    text = clean_text(value).replace(",", "")
    if text.endswith(".0"):
        text = text[:-2]
    if text.isdigit():
        return int(text)
    return None


def mostly_uppercase(text: str) -> bool:
    """Detect category heading rows from the original food workbook."""
    letters = [character for character in text if character.isalpha()]
    if len(letters) < 3:
        return False
    uppercase_count = sum(1 for character in letters if character.isupper())
    return (uppercase_count / len(letters)) >= 0.7


def status_is_allowed(row_status: Any, allowed_statuses: set[str]) -> bool:
    """Compare statuses case-insensitively."""
    return clean_text(row_status).lower() in allowed_statuses


def build_review_maps(
    review_path: Path, allowed_statuses: set[str]
) -> tuple[dict[str, str], dict[tuple[str, str], str], dict[int, str]]:
    """Build lookup maps for category, food group, and food description translations."""
    workbook = pd.read_excel(review_path, sheet_name=None, dtype=object)
    category_map: dict[str, str] = {}
    food_map: dict[tuple[str, str], str] = {}
    description_map: dict[int, str] = {}

    # Category names are unique enough in this dataset to key by English text.
    for _, row in workbook.get("categories", pd.DataFrame()).iterrows():
        if not status_is_allowed(row.get("status"), allowed_statuses):
            continue
        english = clean_text(row.get("english_text"))
        translation = clean_text(row.get("suggested_ro"))
        if english and translation:
            category_map[english] = translation

    # Food group names can repeat across categories, so the category is part of the key.
    for _, row in workbook.get("foods", pd.DataFrame()).iterrows():
        if not status_is_allowed(row.get("status"), allowed_statuses):
            continue
        category = clean_text(row.get("category_en"))
        english = clean_text(row.get("english_text"))
        translation = clean_text(row.get("suggested_ro"))
        if category and english and translation:
            food_map[(category, english)] = translation

    # DA code is the stable identifier for food description rows in FoodsFinal.
    for _, row in workbook.get("food_descriptions", pd.DataFrame()).iterrows():
        if not status_is_allowed(row.get("status"), allowed_statuses):
            continue
        da_code = clean_int(row.get("da_code"))
        translation = clean_text(row.get("suggested_ro"))
        if da_code is not None and translation:
            description_map[da_code] = translation

    return category_map, food_map, description_map


def find_header_column(sheet: Any, header_name: str) -> int | None:
    """Find a header column by exact display name."""
    for cell in sheet[1]:
        if cell.value == header_name:
            return int(cell.column)
    return None


def find_or_create_column(sheet: Any, header_name: str) -> int:
    """Reuse a translation column or create it at the end."""
    existing_column = find_header_column(sheet, header_name)
    if existing_column is not None:
        return existing_column

    next_column = sheet.max_column + 1
    sheet.cell(row=1, column=next_column).value = header_name
    return next_column


def apply_translations(
    *,
    workbook_path: Path,
    output_path: Path,
    review_path: Path,
    allowed_statuses: set[str],
    sheet_name: str | None,
) -> dict[str, int]:
    """Add Romanian translation columns to a copy of FoodsFinal.xlsx."""
    category_map, food_map, description_map = build_review_maps(
        review_path=review_path, allowed_statuses=allowed_statuses
    )

    workbook = load_workbook(workbook_path)
    sheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]

    da_code_column = 1
    description_column = 2
    category_ro_column = find_or_create_column(sheet, "category_ro")
    food_ro_column = find_or_create_column(sheet, "food_ro")
    food_description_ro_column = find_or_create_column(sheet, "food_description_ro")

    # The source workbook is hierarchical, so Food rows inherit the last Category row.
    current_category = ""
    stats = {
        "category_rows_filled": 0,
        "food_rows_filled": 0,
        "food_description_rows_filled": 0,
    }

    for row_index in range(2, sheet.max_row + 1):
        da_code_value = sheet.cell(row=row_index, column=da_code_column).value
        description_value = sheet.cell(row=row_index, column=description_column).value
        da_text = clean_text(da_code_value)
        description = clean_text(description_value)
        da_code = clean_int(da_code_value)

        if da_code is not None:
            translation = description_map.get(da_code)
            if translation:
                sheet.cell(row=row_index, column=food_description_ro_column).value = translation
                stats["food_description_rows_filled"] += 1
            continue

        if da_text and mostly_uppercase(da_text):
            current_category = da_text
            translation = category_map.get(da_text)
            if translation:
                sheet.cell(row=row_index, column=category_ro_column).value = translation
                stats["category_rows_filled"] += 1
            continue

        if description:
            translation = food_map.get((current_category, description))
            if translation:
                sheet.cell(row=row_index, column=food_ro_column).value = translation
                stats["food_rows_filled"] += 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return stats


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Copy reviewed Romanian translations into a FoodsFinal Excel copy."
    )
    parser.add_argument("--review", default=str(DEFAULT_REVIEW), help="Review Excel file.")
    parser.add_argument("--input", default=str(DEFAULT_INPUT), help="FoodsFinal Excel file.")
    parser.add_argument(
        "--output",
        default=str(DEFAULT_OUTPUT),
        help="Translated FoodsFinal Excel file to create.",
    )
    parser.add_argument("--sheet", help="FoodsFinal sheet to update. Defaults to first sheet.")
    parser.add_argument(
        "--statuses",
        nargs="+",
        default=["approved"],
        help="Statuses allowed for export. Example: --statuses approved needs_review",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_arguments()
    review_path = resolve_path(args.review)
    input_path = resolve_path(args.input)
    output_path = resolve_path(args.output)

    if not review_path.exists():
        print(f"ERROR: Review file not found: {review_path}")
        return 1
    if not input_path.exists():
        print(f"ERROR: FoodsFinal file not found: {input_path}")
        return 1

    allowed_statuses = {status.lower() for status in args.statuses}
    stats = apply_translations(
        workbook_path=input_path,
        output_path=output_path,
        review_path=review_path,
        allowed_statuses=allowed_statuses,
        sheet_name=args.sheet,
    )

    print(f"Review input: {review_path}")
    print(f"Foods input:  {input_path}")
    print(f"Output:       {output_path}")
    print(f"Statuses:     {', '.join(sorted(allowed_statuses))}")
    print(f"Category rows filled:         {stats['category_rows_filled']}")
    print(f"Food rows filled:             {stats['food_rows_filled']}")
    print(f"Food description rows filled: {stats['food_description_rows_filled']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
